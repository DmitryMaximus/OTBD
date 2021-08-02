import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import  os
from PyQt5.Qt import QMessageBox
from PyQt5 import QtWidgets
from Progress import ProgressBar
def aggregation(data):
    """
    Логика агрегации родительских и дочерних строк модели tree_model
    Итерация по каждой родительской строке, если запись имеет дочернюю запись:
    значение поля "Наименование главное рус"  дочерней строки добавляется в поле "Альтернативное наименование 1..5" если хотя бы одно поле пустое и
    значение не пристутствует в одном из полей 1-5
    значение поля из списка_1 дочерней строки добавляется в аналогичное поле родительской через "," если такого значения еще нет в род. поле
    список_1 = ["Страна ограничений","Программа ограничений","Связь с санкционным элементом (наим)","Связь с санкционным элементом (наим лат)",
    "INN материнской компании","ОГРН материнской компании"]
    """
    cols = [data.headerData(i, 1) for i in range(0, data.columnCount())]
    df_parent = pd.DataFrame([], columns=cols)
    df_child = pd.DataFrame([], columns=cols+["parent_id"])
    for column in range(0,data.columnCount()):
        for row in range(0,  data.rowCount()):
            df_parent.at[row, cols[column]] = data.item(row, column).text()

    new_child_row = 0
    for row in range(0,  data.rowCount()):
        if data.item(row,0).hasChildren():
            i=0
            while data.item(row, 0).child(i,0) != None:
                for col in range(0,data.columnCount()):
                    df_child.at[new_child_row, cols[col]] = data.item(row, 0).child(i, col).text() if not pd.isna(data.item(row, 0).child(i, col)) else ''
                df_child.at[new_child_row, ["parent_id"]] = data.item(row, 0).text() if not pd.isna(data.item(row, 0)) else ''
                new_child_row = len(df_child) + 1
                i+=1

    list_del = ["Страна ограничений","Программа ограничений","Связь с санкционным элементом (наим)",
                "Связь с санкционным элементом (наим лат)","INN материнской компании","ОГРН материнской компании", "ПИН клиента"]
    for i, row in df_parent.iterrows():
        if len(df_child[df_child["parent_id"]==row["id"]])>0:
            for col in list_del:
                new_val = " | ".join(list(set(df_child[df_child["parent_id"] == row["id"]][col]))+[row[col]])
                df_parent.at[i,col] =  new_val if not new_val.strip() == "|" else ""

            alt_names = list(set(list(set(row[["Альтернативные наименования для загрузки "+str(i) for i in range(1,6)]]))+df_child[df_child["parent_id"] == row["id"]]['Наименование(главное рус)'].values))[:4]
            for idx in range(0,len(alt_names)):
                df_parent.at[i,"Альтернативные наименования для загрузки "+str(idx+1)] = alt_names[idx] if not alt_names[idx] == row["Наименование(главное рус)"] else ""

            new_codes = " , ".join(list(list(set(df_child[df_child["parent_id"] == row["id"]]["Коды санкционных ограничений"]))+[row["Коды санкционных ограничений"]]))
            df_parent.at[i, "Коды санкционных ограничений"] = new_codes.strip() if not len(new_codes.strip()) == 1 else ""
    return df_parent

def create_version(data):
        ret = QMessageBox.question(QtWidgets.QWidget(), '',
                                   "Хотите ли вы сделать текущую версию списка эталонной?",
                                   QMessageBox.Yes | QMessageBox.No)

        if ret == QMessageBox.Yes:
            data.to_excel(os.path.dirname(os.path.abspath(__file__)) + "\exmp.xlsx",index=False)
        else:
            pass


def create_delta(data,path_to_save):
    path_to_save = path_to_save + "/Delta.xlsx"
    df_main = aggregation(data).fillna("")
    create_version(df_main)
    df_delta = pd.read_excel(os.path.dirname(os.path.abspath(__file__))+ "\exmp.xlsx",dtype=str).fillna("")
    df_delta['Комментарий_дельта'] = ""
    df_delta['col_num'] = ""
    Progress = ProgressBar(len(df_delta))
    for i, row in df_delta.iterrows():
        if row['Тип записи'] in ['ЮЛ','0',0]:
            checker = df_main['ИНН']==row['ИНН']
            if len(df_delta[df_delta['ИНН']==row['ИНН']])>1:
                df_delta.at[i,'Комментарий_дельта'] = 'Запись дублируется/'

            if len(df_main[checker])==0:
                df_delta.at[i, 'Комментарий_дельта'] += "Запись на удаление"

            elif len(df_main[checker])>0:
                new_row = df_main[checker].iloc[0]
                for col in list(df_main):
                    if col in list(df_delta) and new_row[col] != row[col]:
                        df_delta.at[i, 'col_num'] += "|"+str(list(df_delta).index(col))
                        df_delta.at[i, col] = new_row[col]
                df_delta.at[i, 'Комментарий_дельта'] += "В записи присутствуют изменения"
            Progress.SendStep()
        elif row['Тип записи'] in ['ФЛ',1,'1']:
            checker = (df_main['ФИО']==row['ФИО'])&(df_main['Дата рождения']==row['Дата рождения'])
            if len(df_delta[df_delta['ИНН'] == row['ИНН']]) > 1:
                df_delta.at[i, 'Комментарий_дельта'] = 'Запись дублируется/'

            if len(df_main[checker]) == 0:
                df_delta.at[i, 'Комментарий_дельта'] += "Запись на удаление"

            elif len(df_main[checker]) > 0:
                new_row = df_main[checker].iloc[0]
                for col in list(df_main):
                    if col in list(df_delta) and new_row[col] != row[col]:
                        df_delta.at[i, 'col_num'] += "|" + str(list(df_delta).index(col))
                        df_delta.at[i, col] = new_row[col]
                df_delta.at[i, 'Комментарий_дельта'] += "В записи присутствуют изменения"
            Progress.SendStep()

    for i, row in df_main.iterrows():
        if len(df_delta[df_delta['ИНН']==row['ИНН']]) == 0 or len(df_delta[(df_delta['ФИО']==row['ФИО'])&(df_delta['Дата рождения']==row['Дата рождения'])]) == 0:
            row_num = len(df_delta) + 1
            for col in list(df_main):
                if col in list(df_delta):
                    df_delta.at[row_num,col] = row[col]
                    df_delta.at[row_num, 'Комментарий_дельта'] = 'Добавлена новая запись'

    cols_change = df_delta['col_num']
    df_delta = df_delta.drop(columns = ['col_num'])
    df_delta.to_excel(path_to_save, index=False)

    workbook = openpyxl.load_workbook(path_to_save)
    workbook.active
    sheet = workbook['Sheet1']
    idx=1
    for items in cols_change:
        if items != "" and not pd.isna(items):
            items = items[1:].split("|")
            for col in list(map(int,items)):
                cell_path = sheet.cell(idx+1, col+1)
                cell_path.fill = PatternFill(start_color="00FFCC00",
                                                                           end_color="00FFCC00", fill_type='solid')
        idx+=1
    workbook.save(path_to_save)
    workbook.close()